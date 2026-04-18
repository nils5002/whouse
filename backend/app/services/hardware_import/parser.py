from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from .types import CANONICAL_COLUMNS, ParsedExcelFile, ParsedExcelRow, REQUIRED_COLUMNS

SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm"}

COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "name": (
        "name",
        "geraetename",
        "gerätename",
        "bezeichnung",
        "gerät",
        "geraet",
        "device",
    ),
    "inventory_number": (
        "nummer",
        "inventarnummer",
        "inventar nummer",
        "inventory number",
        "asset number",
        "ipad",
    ),
    "model": ("modell", "model", "typ"),
    "description": ("beschreibung", "description", "kommentar", "details", "bemerkung", "notes"),
    "serial_number": (
        "seriennummer",
        "serialnumber",
        "serial",
        "sn",
    ),
    "sim_number": (
        "sim karten nummer",
        "sim-kartennummer",
        "sim kartennummer",
        "sim-karte",
        "sim",
    ),
    "phone_number": ("rufnummer", "telefonnummer", "telefon", "phone", "phone number"),
    "power_supply": ("netzteil", "power supply", "stromversorgung"),
    "language": ("sprache", "language"),
    "ip_address": ("ip", "ipadresse", "ip-adresse", "ip address"),
    "mac_lan": (
        "mac lan",
        "mac_lan",
        "maclan",
        "mac ethernet",
        "ethernet mac",
        "mac-adresse",
        "mac adresse",
        "mac address",
        "mac-adresse lan",
        "mac adresse lan",
    ),
    "mac_wlan": (
        "mac wlan",
        "mac_wlan",
        "macwlan",
        "wifi mac",
        "wlan mac",
        "mac-adresse wlan",
        "mac adresse wlan",
    ),
    "status": ("status", "zustand"),
}


def list_excel_files(import_dir: Path) -> tuple[list[Path], list[str]]:
    files: list[Path] = []
    skipped: list[str] = []
    for path in sorted(import_dir.iterdir(), key=lambda p: p.name.lower()):
        if not path.is_file():
            continue
        if path.suffix.lower() in SUPPORTED_EXTENSIONS:
            files.append(path)
        else:
            skipped.append(path.name)
    return files, skipped


def parse_excel_file(path: Path) -> ParsedExcelFile:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook.active
        parsed = ParsedExcelFile(path=path, file_name=path.name, sheet_name=sheet.title)

        row_iter = sheet.iter_rows(values_only=True)
        buffered_rows = _take_rows(row_iter, limit=30)
        header_index, col_idx_by_canonical = find_header_row(buffered_rows)
        if header_index is None:
            parsed.missing_required_columns = list(REQUIRED_COLUMNS)
            return parsed

        parsed.missing_required_columns = [
            col for col in REQUIRED_COLUMNS if col not in col_idx_by_canonical
        ]
        parsed.missing_optional_columns = [
            col for col in CANONICAL_COLUMNS if col not in col_idx_by_canonical and col not in REQUIRED_COLUMNS
        ]

        header_row_pos = next(
            (idx for idx, (row_number, _) in enumerate(buffered_rows) if row_number == header_index),
            None,
        )
        if header_row_pos is None:
            parsed.missing_required_columns = list(REQUIRED_COLUMNS)
            return parsed

        for row_number, raw_row in buffered_rows[header_row_pos + 1 :]:
            if is_header_repeat(raw_row, col_idx_by_canonical):
                continue
            data: dict[str, object] = {}
            for canonical, idx in col_idx_by_canonical.items():
                value = raw_row[idx] if idx < len(raw_row) else None
                data[canonical] = value
            if any(value not in (None, "") for value in data.values()):
                parsed.rows.append(
                    ParsedExcelRow(
                        file_name=path.name,
                        sheet_name=sheet.title,
                        row_number=row_number,
                        data=data,
                    )
                )

        row_number = (buffered_rows[-1][0] + 1) if buffered_rows else 1
        for raw_row in row_iter:
            if is_header_repeat(raw_row, col_idx_by_canonical):
                row_number += 1
                continue
            data: dict[str, object] = {}
            for canonical, idx in col_idx_by_canonical.items():
                value = raw_row[idx] if idx < len(raw_row) else None
                data[canonical] = value
            if any(value not in (None, "") for value in data.values()):
                parsed.rows.append(
                    ParsedExcelRow(
                        file_name=path.name,
                        sheet_name=sheet.title,
                        row_number=row_number,
                        data=data,
                    )
                )
            row_number += 1
        return parsed
    finally:
        workbook.close()


def normalize_header(value: object) -> str:
    text = str(value or "").strip().lower()
    return " ".join(text.replace("_", " ").replace("-", " ").split())


NORMALIZED_ALIAS_MAP: dict[str, set[str]] = {
    canonical: {normalize_header(alias) for alias in aliases}
    for canonical, aliases in COLUMN_ALIASES.items()
}


def resolve_columns(header: list[str]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for canonical, alias_set in NORMALIZED_ALIAS_MAP.items():
        for idx, header_name in enumerate(header):
            if header_name in alias_set:
                mapping[canonical] = idx
                break
    return mapping


def find_header_row(buffered_rows: list[tuple[int, tuple[object, ...]]]) -> tuple[int | None, dict[str, int]]:
    best_row_index: int | None = None
    best_mapping: dict[str, int] = {}
    best_score = -1
    best_required = -1

    for row_index, row_values in buffered_rows:
        header = [normalize_header(value) for value in row_values]
        if not any(header):
            continue
        mapping = resolve_columns(header)
        score = len(mapping)
        required_matches = sum(1 for col in REQUIRED_COLUMNS if col in mapping)
        if score > best_score or (score == best_score and required_matches > best_required):
            best_score = score
            best_required = required_matches
            best_row_index = row_index
            best_mapping = mapping

    if best_row_index is None or best_score <= 0:
        return None, {}
    return best_row_index, best_mapping


def is_header_repeat(raw_row: tuple[object, ...], col_idx_by_canonical: dict[str, int]) -> bool:
    checks = 0
    matches = 0
    for canonical, idx in col_idx_by_canonical.items():
        value = normalize_header(raw_row[idx] if idx < len(raw_row) else "")
        if not value:
            continue
        checks += 1
        if value in NORMALIZED_ALIAS_MAP[canonical]:
            matches += 1
    return checks > 0 and matches == checks


def _take_rows(
    row_iter: Iterable[tuple[object, ...]],
    *,
    limit: int,
) -> list[tuple[int, tuple[object, ...]]]:
    rows: list[tuple[int, tuple[object, ...]]] = []
    for row_number, row in enumerate(row_iter, start=1):
        rows.append((row_number, row))
        if row_number >= limit:
            break
    return rows
